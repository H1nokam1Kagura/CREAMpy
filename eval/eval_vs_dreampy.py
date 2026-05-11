"""
CREAMpy cross-validation harness against IFPRI DREAMpy v2.2.3.

Two evaluation modes
--------------------
Mode A — Analytical (no DREAMpy required)
    Runs the 9 built-in closed-form test cases and reports pass/fail.

Mode B — Cross-validation against DREAMpy executable
    Generates Excel inputs for four canonical test vectors, runs DREAMpy,
    parses the output workbooks, then compares ΔPS / ΔCS / NPV_W against
    CREAMpy. Acceptable divergence ≤ 2 % (configurable via --tolerance).

Canonical test vectors
----------------------
  TV1  Symmetric elasticities (ε=|η|=0.5), full adoption, single year.
       ΔPS must equal ΔCS; ΔW = K·P0·Q0·(1 + 0.5·K·Z)
  TV2  Inelastic demand (η=−0.2): cs_share > ps_share for K-shift.
  TV3  Elastic demand   (η=−1.5): ps_share > cs_share for K-shift.
  TV4  10-year linear adoption ramp (0.1→1.0), NPV compounded at 5 %.

Usage
-----
  # Mode A only (no DREAMpy needed):
  python eval/eval_vs_dreampy.py

  # Mode B — supply path to DREAMpy executable or batch file:
  python eval/eval_vs_dreampy.py --dreampy-exe /path/to/DREAMpy.bat

  # Adjust tolerance (default 2 %):
  python eval/eval_vs_dreampy.py --dreampy-exe ... --tolerance 0.01

  # Write report to file:
  python eval/eval_vs_dreampy.py --dreampy-exe ... --report eval_report.md

DREAMpy download
----------------
  DREAMpy v2.2.3 is available from IFPRI:
  https://www.ifpri.org/project/dream

Notes on expected divergence
-----------------------------
  Sources of divergence ≤ 2 % (all acceptable):
    * Python float64 vs Excel double intermediate rounding    (< 0.001 %)
    * Excel cell rounding display vs stored value            (< 0.01  %)
    * Adoption curve discretisation differences              (< 0.5   %)

  Divergence > 2 % indicates a substantive formula difference:
    * DREAMpy may fold ptrs into K (risk double-counted here vs there)
    * DREAMpy may apply a slightly different second-order correction
    * Template auto-fill missed a cell — see printed cell map
"""

from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import sys
import textwrap
from datetime import datetime
from pathlib import Path
from typing import Optional

# ── Test vectors (shared by both modes) ──────────────────────────────────────

TEST_VECTORS = [
    dict(
        id="TV1",
        description="Symmetric elasticities, full adoption, single year",
        K=0.15, epsilon=0.5, eta=-0.5,
        P0=200.0, Q0=1_000_000.0,
        years=[2024], adoption_fracs=[1.0],
        base_year=2024, discount_rate=0.05,
        note="ΔPS must equal ΔCS; ΔW ≈ 30,562,500",
    ),
    dict(
        id="TV2",
        description="Inelastic demand: cs_share > ps_share",
        K=0.10, epsilon=0.5, eta=-0.2,
        P0=150.0, Q0=800_000.0,
        years=[2024], adoption_fracs=[1.0],
        base_year=2024, discount_rate=0.05,
        note="K-shift: cs_share = ε/(ε+|η|) = 5/7 ≈ 0.714",
    ),
    dict(
        id="TV3",
        description="Elastic demand: ps_share > cs_share",
        K=0.10, epsilon=0.5, eta=-1.5,
        P0=150.0, Q0=800_000.0,
        years=[2024], adoption_fracs=[1.0],
        base_year=2024, discount_rate=0.05,
        note="K-shift: ps_share = |η|/(ε+|η|) = 1.5/2.0 = 0.75",
    ),
    dict(
        id="TV4",
        description="10-year adoption ramp 0.1→1.0, NPV at 5 %",
        K=0.15, epsilon=0.5, eta=-0.5,
        P0=200.0, Q0=1_000_000.0,
        years=list(range(2025, 2035)),
        adoption_fracs=[0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0],
        base_year=2025, discount_rate=0.05,
        note="NPV_W ≈ 126,352,930",
    ),
]


# ── Mode A: analytical validation ────────────────────────────────────────────

def run_mode_a() -> dict:
    """Run built-in 9-case validation. Returns {passed, failed}."""
    sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
    from creampy.__main__ import run_validation
    ok = run_validation()
    return {"passed": ok, "failed_count": 0 if ok else 1}


# ── Mode B helpers ────────────────────────────────────────────────────────────

def _creampy_results(tv: dict) -> dict:
    """Run CREAMpy on a test vector and return NPV results."""
    sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
    from creampy import ClosedEconomy, ModelParams
    p = ModelParams(
        K=tv["K"], epsilon=tv["epsilon"], eta=tv["eta"],
        P0=tv["P0"], Q0=tv["Q0"],
        years=tv["years"], adoption_fracs=tv["adoption_fracs"],
        base_year=tv["base_year"], discount_rate=tv["discount_rate"],
    )
    r = ClosedEconomy(p).run()
    return {
        "npv_W":  r.npv_W,
        "npv_PS": r.npv_PS,
        "npv_CS": r.npv_CS,
        "year_results": [
            {"year": yr.year, "dPS": yr.dPS, "dCS": yr.dCS, "dW": yr.dW}
            for yr in r.year_results
        ],
    }


def _fill_dreampy_template(template_path: Path, tv: dict, out_path: Path) -> list[str]:
    """
    Copy template and attempt to fill known parameter cells.
    Returns list of cells filled (for diagnostics).

    Template auto-fill strategy: scan all sheets for label cells whose text
    contains keywords, then write the parameter value into the adjacent cell.
    Falls back gracefully when openpyxl is not installed or labels are not found.
    """
    shutil.copy(template_path, out_path)

    try:
        import openpyxl
    except ImportError:
        print("  [warn] openpyxl not installed — skipping template fill")
        print("         pip install openpyxl  then re-run for Mode B")
        return []

    wb = openpyxl.load_workbook(out_path)
    fills: list[str] = []

    FIELD_MAP = {
        ("supply", "elast"):  tv["epsilon"],
        ("demand", "elast"):  tv["eta"],
        ("price",):           tv["P0"],
        ("quantity",):        tv["Q0"],
        ("shift",):           tv["K"],
        ("discount",):        tv["discount_rate"],
        ("base", "year"):     tv["base_year"],
    }

    def find_adjacent(ws, *keywords: str):
        kl = [k.lower() for k in keywords]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cv = cell.value.lower()
                    if all(k in cv for k in kl):
                        nc = ws.cell(row=cell.row, column=cell.column + 1)
                        return nc
        return None

    for ws in wb.worksheets:
        for keywords, value in FIELD_MAP.items():
            c = find_adjacent(ws, *keywords)
            if c:
                c.value = value
                fills.append(f"{ws.title}!{c.coordinate}={value}")

    wb.save(out_path)
    return fills


def _parse_dreampy_output(output_dir: Path) -> Optional[dict]:
    """
    Parse DREAMpy output Excel files and extract NPV metrics.
    Returns dict with npv_W/npv_PS/npv_CS, or None if parsing fails.
    """
    try:
        import openpyxl
    except ImportError:
        return None

    xlsx_files = list(output_dir.glob("*.xlsx"))
    if not xlsx_files:
        return None

    extracted: dict = {}
    wb = openpyxl.load_workbook(xlsx_files[0], data_only=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                cv = cell.value.lower()
                adj = ws.cell(row=cell.row, column=cell.column + 1).value
                if not isinstance(adj, (int, float)):
                    continue
                v = float(adj)
                if "npv" in cv and "producer" in cv:   extracted["npv_PS"] = v
                elif "npv" in cv and "consumer" in cv:  extracted["npv_CS"] = v
                elif "npv" in cv and ("total" in cv or "welfare" in cv): extracted["npv_W"] = v

    return extracted if extracted else None


def run_mode_b(dreampy_exe: Path, work_dir: Path, tolerance: float) -> list[dict]:
    """Cross-validate CREAMpy against DREAMpy for all test vectors."""

    # Find Excel template inside DREAMpy distribution
    template: Optional[Path] = None
    for pattern in ["*ClosedEconomy*.xlsx", "*closed*.xlsx", "*.xlsx"]:
        hits = list(dreampy_exe.parent.rglob(pattern))
        if hits:
            template = hits[0]
            break
    if template is None:
        print("[warn] Could not locate a Closed Economy Excel template in DREAMpy directory.")
        print("       Pass the template path manually or fill inputs by hand.")

    comparisons = []
    for tv in TEST_VECTORS:
        tv_id = tv["id"]
        print(f"\n── {tv_id}: {tv['description']}")

        # CREAMpy reference result
        cream = _creampy_results(tv)
        print(f"   CREAMpy  NPV_W = {cream['npv_W']:>18,.0f}")

        # Generate DREAMpy input
        tv_dir = work_dir / tv_id
        tv_dir.mkdir(parents=True, exist_ok=True)
        input_path = tv_dir / f"{tv_id}_input.xlsx"

        if template:
            fills = _fill_dreampy_template(template, tv, input_path)
            print(f"   Filled {len(fills)} template cells: {fills[:3]}")
        else:
            print(f"   [skip] No template — skipping DREAMpy run for {tv_id}")
            comparisons.append({"tv": tv_id, "status": "skip", "reason": "no_template"})
            continue

        # Run DREAMpy
        out_dir = tv_dir / "dreampy_out"
        out_dir.mkdir(exist_ok=True)
        try:
            proc = subprocess.run(
                [str(dreampy_exe)],
                input=f"{input_path}\n{out_dir}\n",
                capture_output=True, text=True, timeout=120,
            )
            print(f"   DREAMpy exit code: {proc.returncode}")
        except subprocess.TimeoutExpired:
            print(f"   [error] DREAMpy timed out for {tv_id}")
            comparisons.append({"tv": tv_id, "status": "error", "reason": "timeout"})
            continue
        except Exception as exc:
            print(f"   [error] DREAMpy failed: {exc}")
            comparisons.append({"tv": tv_id, "status": "error", "reason": str(exc)})
            continue

        # Parse output
        dream = _parse_dreampy_output(out_dir)
        if not dream:
            print(f"   [warn] Could not parse DREAMpy output for {tv_id}")
            comparisons.append({"tv": tv_id, "status": "parse_failed"})
            continue

        print(f"   DREAMpy  NPV_W = {dream.get('npv_W', float('nan')):>18,.0f}")

        # Compare
        row = {"tv": tv_id, "metrics": []}
        all_pass = True
        for metric, ck, dk in [("NPV_W", "npv_W", "npv_W"),
                                ("NPV_PS", "npv_PS", "npv_PS"),
                                ("NPV_CS", "npv_CS", "npv_CS")]:
            cv = cream.get(ck)
            dv = dream.get(dk)
            if cv is None or dv is None:
                continue
            rel = abs(cv - dv) / max(abs(dv), 1.0)
            ok  = rel <= tolerance
            if not ok:
                all_pass = False
            row["metrics"].append({
                "metric": metric, "creampy": cv, "dreampy": dv,
                "rel_diff": rel, "pass": ok,
            })
            flag = "PASS" if ok else "FAIL"
            print(f"   {metric:8s}  rel_diff={rel*100:.3f}%  [{flag}]")

        row["status"] = "pass" if all_pass else "fail"
        comparisons.append(row)

    return comparisons


# ── Report ────────────────────────────────────────────────────────────────────

def write_report(mode_a: dict, mode_b: Optional[list], tolerance: float,
                 report_path: Path) -> None:
    lines = [
        "# CREAMpy cross-validation report",
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        f"Tolerance: {tolerance * 100:.1f} %",
        "",
        "## Mode A — Analytical validation",
        "",
        "Nine closed-form cases with exact solutions (zero inputs, symmetric and",
        "asymmetric elasticities, K-shift vs J-shift partition, single-year NPV,",
        "two-year NPV, linearity in P0 and Q0).",
        "",
        f"> **{'PASS' if mode_a['passed'] else 'FAIL'}** — "
        f"{'all 9 cases match analytical solution' if mode_a['passed'] else 'failures detected'}.",
        "",
        "## Mode B — Cross-validation against DREAMpy v2.2.3",
        "",
    ]

    if mode_b is None:
        lines += [
            "> **Skipped.** `--dreampy-exe` not provided.",
            ">",
            "> Download DREAMpy from https://www.ifpri.org/project/dream",
            "> then re-run with `--dreampy-exe /path/to/DREAMpy.bat`.",
        ]
    else:
        lines += [
            "| TV | Description | NPV_W (CREAMpy) | NPV_W (DREAMpy) | Δ% | Result |",
            "|----|-------------|-----------------|-----------------|-----|--------|",
        ]
        for r in mode_b:
            if r["status"] in ("skip", "error", "parse_failed"):
                lines.append(f"| {r['tv']} | — | — | — | — | {r['status'].upper()} |")
                continue
            npv_m = next((m for m in r["metrics"] if m["metric"] == "NPV_W"), None)
            if npv_m:
                tv = next(t for t in TEST_VECTORS if t["id"] == r["tv"])
                lines.append(
                    f"| {r['tv']} | {tv['description'][:40]} "
                    f"| {npv_m['creampy']:,.0f} | {npv_m['dreampy']:,.0f} "
                    f"| {npv_m['rel_diff']*100:.2f}% | {'PASS' if r['status']=='pass' else 'FAIL'} |"
                )

    lines += [
        "",
        "## Expected sources of divergence",
        "",
        "All of the following are acceptable at ≤ 2 %:",
        "- Python float64 vs Excel double intermediate rounding (< 0.001 %)",
        "- Excel cell display rounding (< 0.01 %)",
        "- Adoption curve discretisation differences (< 0.5 %)",
        "",
        "Divergence > 2 % requires investigation:",
        "- DREAMpy may apply ptrs inside K (risk double-counted relative to CREAMpy)",
        "- DREAMpy may use a slightly different second-order correction term",
        "- Template auto-fill may have missed a cell — inspect cell map output",
        "",
        "---",
        "_Generated by CREAMpy eval/eval_vs_dreampy.py_",
    ]

    report_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"\nReport written to {report_path}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--dreampy-exe", default=None,
                   help="Path to DREAMpy executable or batch file (enables Mode B).")
    p.add_argument("--work-dir", default=None,
                   help="Directory for intermediate files (default: ./dreampy_eval_<ts>).")
    p.add_argument("--tolerance", type=float, default=0.02,
                   help="Acceptable relative divergence for Mode B (default 0.02 = 2%%).")
    p.add_argument("--report", default=None,
                   help="Write Markdown report to this path.")
    args = p.parse_args()

    print("=== CREAMpy evaluation harness ===\n")

    # Mode A
    print("── Mode A: analytical validation")
    mode_a = run_mode_a()

    # Mode B
    mode_b = None
    if args.dreampy_exe:
        exe = Path(args.dreampy_exe)
        if not exe.exists():
            print(f"[error] DREAMpy executable not found: {exe}")
            return 1
        ts = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
        work = Path(args.work_dir) if args.work_dir else Path(f"dreampy_eval_{ts}")
        work.mkdir(parents=True, exist_ok=True)
        print(f"\n── Mode B: cross-validation against DREAMpy")
        print(f"   exe:      {exe}")
        print(f"   work dir: {work}")
        print(f"   tolerance: {args.tolerance * 100:.1f} %")
        mode_b = run_mode_b(exe, work, args.tolerance)
    else:
        print("\n── Mode B: skipped (pass --dreampy-exe to enable)")

    # Report
    if args.report:
        write_report(mode_a, mode_b, args.tolerance, Path(args.report))

    # Summary
    print("\n=== Summary ===")
    print(f"Mode A: {'PASS' if mode_a['passed'] else 'FAIL'}")
    if mode_b is not None:
        passes = sum(1 for r in mode_b if r.get("status") == "pass")
        total  = len(mode_b)
        print(f"Mode B: {passes}/{total} test vectors PASS")

    return 0 if mode_a["passed"] else 1


if __name__ == "__main__":
    sys.exit(main())
